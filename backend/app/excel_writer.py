"""
Excel writer for continuous daily trade logging.

This module provides thread-safe Excel file writing with:
- Daily file rotation (one file per day: trades_YYYYMMDD.xlsx)
- Three sheets: Trades, Strategies, Analytics
- Background thread for asynchronous writes
- Duplicate prevention (updates existing trades instead of creating duplicates)
- Trade loading on startup (for state persistence)

The ExcelWriter uses a queue-based architecture where write operations
are queued and processed by a background thread, ensuring non-blocking
writes and thread safety.
"""

import asyncio
import logging
from datetime import datetime, date
from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import threading
from queue import Queue, Empty

from app.config import EXCEL_OUTPUT_DIR
from app.models import Trade, Strategy, Analytics

logger = logging.getLogger(__name__)


class ExcelWriter:
    """
    Thread-safe Excel writer with daily file rotation.
    
    This class manages continuous writing to Excel files with the following features:
    - Daily file rotation (creates new file each day)
    - Thread-safe writes via background thread and queue
    - Duplicate prevention (updates existing trades by ID)
    - Automatic file loading on startup
    - Three sheets: Trades, Strategies, Analytics
    
    The writer uses a queue-based architecture where write operations are
    queued and processed asynchronously by a background thread.
    
    Attributes:
        current_date: Current date for file naming
        write_queue: Queue for write operations (thread-safe)
        lock: Thread lock for critical sections
        workbook: Current openpyxl Workbook instance
        trades_sheet: Worksheet for trades
        strategies_sheet: Worksheet for strategies
        analytics_sheet: Worksheet for analytics
        current_file_path: Path to current Excel file
        running: Flag to control background thread
        writer_thread: Background thread for processing writes
    """
    
    def __init__(self):
        self.current_date = date.today()
        self.write_queue = Queue()
        self.lock = threading.Lock()
        self.workbook: Optional[Workbook] = None
        self.trades_sheet = None
        self.strategies_sheet = None
        self.analytics_sheet = None
        self.current_file_path: Optional[Path] = None
        self.running = True
        
        # Start background writer thread
        self.writer_thread = threading.Thread(target=self._writer_loop, daemon=True)
        self.writer_thread.start()
        
        # Initialize today's file
        self._ensure_file_exists()
    
    def _get_file_path(self, target_date: date) -> Path:
        """Get Excel file path for a given date."""
        filename = f"trades_{target_date.strftime('%Y%m%d')}.xlsx"
        return EXCEL_OUTPUT_DIR / filename
    
    def _ensure_file_exists(self):
        """Create or load today's Excel file."""
        today = date.today()
        
        if today != self.current_date:
            # Date changed - create new file
            self.current_date = today
            self._create_new_file()
        elif self.current_file_path is None or not self.current_file_path.exists():
            # File doesn't exist - create it
            self._create_new_file()
        else:
            # Load existing file
            try:
                self.workbook = load_workbook(self.current_file_path)
                self.trades_sheet = self.workbook["Trades"]
                if "Strategies" not in self.workbook.sheetnames:
                    self.strategies_sheet = self.workbook.create_sheet("Strategies")
                    self._init_strategies_sheet()
                else:
                    self.strategies_sheet = self.workbook["Strategies"]
                if "Analytics" not in self.workbook.sheetnames:
                    self.analytics_sheet = self.workbook.create_sheet("Analytics")
                    self._init_analytics_sheet()
                else:
                    self.analytics_sheet = self.workbook["Analytics"]
            except Exception as e:
                logger.error(f"Error loading Excel file: {e}")
                self._create_new_file()
    
    def _create_new_file(self):
        """Create a new Excel file with headers."""
        self.current_file_path = self._get_file_path(self.current_date)
        self.workbook = Workbook()
        
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Create Trades sheet
        self.trades_sheet = self.workbook.create_sheet("Trades")
        self._init_trades_sheet()
        
        # Create Strategies sheet
        self.strategies_sheet = self.workbook.create_sheet("Strategies")
        self._init_strategies_sheet()
        
        # Create Analytics sheet
        self.analytics_sheet = self.workbook.create_sheet("Analytics")
        self._init_analytics_sheet()
        
        self.workbook.save(self.current_file_path)
        logger.info(f"Created new Excel file: {self.current_file_path}")
    
    def _init_trades_sheet(self):
        """Initialize Trades sheet headers."""
        headers = [
            "ID", "Timestamp", "Action", "Underlying", "Notional Leg1", "Notional Leg2",
            "Currency Leg1", "Currency Leg2", "Fixed Rate Leg1", "Fixed Rate Leg2",
            "Spread Leg2", "Rate %", "Maturity", "Instrument", "Platform", "Strategy ID",
            "Package", "Notional EUR", "Is Forward", "Effective Date"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.trades_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            self.trades_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def _init_strategies_sheet(self):
        """Initialize Strategies sheet headers."""
        headers = [
            "Strategy ID", "Type", "Underlying", "Nb Legs", "Total Notional EUR",
            "Execution Start", "Execution End", "Package Price"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.strategies_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        for col in range(1, len(headers) + 1):
            self.strategies_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    def _init_analytics_sheet(self):
        """Initialize Analytics sheet."""
        self.analytics_sheet.cell(row=1, column=1, value="Metric")
        self.analytics_sheet.cell(row=1, column=2, value="Value")
        self.analytics_sheet.cell(row=1, column=1).font = Font(bold=True)
        self.analytics_sheet.cell(row=1, column=2).font = Font(bold=True)
    
    def append_trade(self, trade: Trade):
        """Queue a trade for writing."""
        self.write_queue.put(("trade", trade))
    
    def update_strategy(self, strategy: Strategy):
        """Queue a strategy update."""
        self.write_queue.put(("strategy", strategy))
    
    def update_analytics(self, analytics: Analytics):
        """Queue analytics update."""
        self.write_queue.put(("analytics", analytics))
    
    def _writer_loop(self):
        """Background thread loop for writing to Excel."""
        while self.running:
            try:
                item = self.write_queue.get(timeout=1)
                item_type, data = item
                
                with self.lock:
                    self._ensure_file_exists()
                    
                    if item_type == "trade":
                        self._write_trade(data)
                    elif item_type == "strategy":
                        self._write_strategy(data)
                    elif item_type == "analytics":
                        self._write_analytics(data)
                    
                    self.workbook.save(self.current_file_path)
                    
            except Empty:
                # Queue is empty, continue waiting - this is normal
                continue
            except Exception as e:
                if self.running:  # Only log if still running
                    logger.error(f"Error in Excel writer loop: {e}", exc_info=True)
    
    def _write_trade(self, trade: Trade):
        """Write a trade to the Trades sheet."""
        # Check if trade already exists (by ID)
        for row in range(2, self.trades_sheet.max_row + 1):
            if self.trades_sheet.cell(row=row, column=1).value == trade.dissemination_identifier:
                # Update existing row
                self.trades_sheet.cell(row=row, column=2, value=trade.execution_timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                self.trades_sheet.cell(row=row, column=3, value=trade.action_type)
                self.trades_sheet.cell(row=row, column=4, value=trade.unique_product_identifier_underlier_name or "")
                self.trades_sheet.cell(row=row, column=5, value=trade.notional_amount_leg1)
                self.trades_sheet.cell(row=row, column=6, value=trade.notional_amount_leg2)
                self.trades_sheet.cell(row=row, column=7, value=trade.notional_currency_leg1)
                self.trades_sheet.cell(row=row, column=8, value=trade.notional_currency_leg2)
                self.trades_sheet.cell(row=row, column=9, value=trade.fixed_rate_leg1)
                self.trades_sheet.cell(row=row, column=10, value=trade.fixed_rate_leg2)
                self.trades_sheet.cell(row=row, column=11, value=trade.spread_leg2)
                rate_display = ""
                if trade.fixed_rate_leg1:
                    rate_display = f"{trade.fixed_rate_leg1 * 100:.4f}%"
                elif trade.spread_leg2 is not None:
                    rate_display = f"Spread: {trade.spread_leg2}"
                self.trades_sheet.cell(row=row, column=12, value=rate_display)
                self.trades_sheet.cell(row=row, column=13, value=trade.expiration_date or "")
                self.trades_sheet.cell(row=row, column=14, value=trade.instrument or "")
                self.trades_sheet.cell(row=row, column=15, value=trade.platform_identifier or "")
                self.trades_sheet.cell(row=row, column=16, value=trade.strategy_id or "")
                self.trades_sheet.cell(row=row, column=17, value="Yes" if trade.package_indicator else "No")
                self.trades_sheet.cell(row=row, column=18, value=trade.notional_eur or 0)
                self.trades_sheet.cell(row=row, column=19, value="Yes" if trade.is_forward else "No")
                self.trades_sheet.cell(row=row, column=20, value=trade.effective_date or "")
                return
        
        # Add new trade
        row = self.trades_sheet.max_row + 1
        
        # Calculate rate display
        rate_display = ""
        if trade.fixed_rate_leg1:
            rate_display = f"{trade.fixed_rate_leg1 * 100:.4f}%"
        elif trade.spread_leg2 is not None:
            rate_display = f"Spread: {trade.spread_leg2}"
        
        self.trades_sheet.cell(row=row, column=1, value=trade.dissemination_identifier)
        self.trades_sheet.cell(row=row, column=2, value=trade.execution_timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        self.trades_sheet.cell(row=row, column=3, value=trade.action_type)
        self.trades_sheet.cell(row=row, column=4, value=trade.unique_product_identifier_underlier_name or "")
        self.trades_sheet.cell(row=row, column=5, value=trade.notional_amount_leg1)
        self.trades_sheet.cell(row=row, column=6, value=trade.notional_amount_leg2)
        self.trades_sheet.cell(row=row, column=7, value=trade.notional_currency_leg1)
        self.trades_sheet.cell(row=row, column=8, value=trade.notional_currency_leg2)
        self.trades_sheet.cell(row=row, column=9, value=trade.fixed_rate_leg1)
        self.trades_sheet.cell(row=row, column=10, value=trade.fixed_rate_leg2)
        self.trades_sheet.cell(row=row, column=11, value=trade.spread_leg2)
        self.trades_sheet.cell(row=row, column=12, value=rate_display)
        self.trades_sheet.cell(row=row, column=13, value=trade.expiration_date or "")
        self.trades_sheet.cell(row=row, column=14, value=trade.instrument or "")
        self.trades_sheet.cell(row=row, column=15, value=trade.platform_identifier or "")
        self.trades_sheet.cell(row=row, column=16, value=trade.strategy_id or "")
        self.trades_sheet.cell(row=row, column=17, value="Yes" if trade.package_indicator else "No")
        self.trades_sheet.cell(row=row, column=18, value=trade.notional_eur or 0)
        self.trades_sheet.cell(row=row, column=19, value="Yes" if trade.is_forward else "No")
        self.trades_sheet.cell(row=row, column=20, value=trade.effective_date or "")
    
    def _write_strategy(self, strategy: Strategy):
        """Write or update a strategy in the Strategies sheet."""
        # Check if strategy already exists
        for row in range(2, self.strategies_sheet.max_row + 1):
            if self.strategies_sheet.cell(row=row, column=1).value == strategy.strategy_id:
                # Update existing row
                self.strategies_sheet.cell(row=row, column=2, value=strategy.strategy_type)
                self.strategies_sheet.cell(row=row, column=3, value=strategy.underlying_name)
                self.strategies_sheet.cell(row=row, column=4, value=len(strategy.legs))
                self.strategies_sheet.cell(row=row, column=5, value=strategy.total_notional_eur)
                self.strategies_sheet.cell(row=row, column=6, value=strategy.execution_start.strftime("%Y-%m-%d %H:%M:%S"))
                self.strategies_sheet.cell(row=row, column=7, value=strategy.execution_end.strftime("%Y-%m-%d %H:%M:%S"))
                self.strategies_sheet.cell(row=row, column=8, value=strategy.package_transaction_price or "")
                return
        
        # Add new strategy
        row = self.strategies_sheet.max_row + 1
        self.strategies_sheet.cell(row=row, column=1, value=strategy.strategy_id)
        self.strategies_sheet.cell(row=row, column=2, value=strategy.strategy_type)
        self.strategies_sheet.cell(row=row, column=3, value=strategy.underlying_name)
        self.strategies_sheet.cell(row=row, column=4, value=len(strategy.legs))
        self.strategies_sheet.cell(row=row, column=5, value=strategy.total_notional_eur)
        self.strategies_sheet.cell(row=row, column=6, value=strategy.execution_start.strftime("%Y-%m-%d %H:%M:%S"))
        self.strategies_sheet.cell(row=row, column=7, value=strategy.execution_end.strftime("%Y-%m-%d %H:%M:%S"))
        self.strategies_sheet.cell(row=row, column=8, value=strategy.package_transaction_price or "")
    
    def _write_analytics(self, analytics: Analytics):
        """Write analytics summary to Analytics sheet."""
        # Clear existing data (keep headers)
        for row in range(2, self.analytics_sheet.max_row + 1):
            for col in range(1, 3):
                self.analytics_sheet.cell(row=row, column=col).value = None
        
        row = 2
        self.analytics_sheet.cell(row=row, column=1, value="Total Trades")
        self.analytics_sheet.cell(row=row, column=2, value=analytics.total_trades)
        row += 1
        
        self.analytics_sheet.cell(row=row, column=1, value="Total Notional EUR")
        self.analytics_sheet.cell(row=row, column=2, value=analytics.total_notional_eur)
        row += 1
        
        self.analytics_sheet.cell(row=row, column=1, value="Average Size EUR")
        self.analytics_sheet.cell(row=row, column=2, value=analytics.avg_size_eur)
        row += 1
        
        self.analytics_sheet.cell(row=row, column=1, value="Largest Trade EUR")
        self.analytics_sheet.cell(row=row, column=2, value=analytics.largest_trade_eur)
        row += 1
        
        self.analytics_sheet.cell(row=row, column=1, value="Strategies Count")
        self.analytics_sheet.cell(row=row, column=2, value=analytics.strategies_count)
        row += 1
        
        # Top underlyings
        row += 1
        self.analytics_sheet.cell(row=row, column=1, value="Top Underlyings")
        row += 1
        for underlying in analytics.top_underlyings[:10]:
            self.analytics_sheet.cell(row=row, column=1, value=underlying["name"])
            self.analytics_sheet.cell(row=row, column=2, value=underlying["notional"])
            row += 1
        
        # Advanced metrics
        if analytics.curve_metrics:
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="=== CURVE METRICS ===")
            row += 1
            for instrument_data in analytics.curve_metrics.instrument_distribution:
                self.analytics_sheet.cell(row=row, column=1, value=f"Instrument {instrument_data['instrument']}")
                self.analytics_sheet.cell(row=row, column=2, value=f"Notional: {instrument_data['notional']}, Count: {instrument_data['count']}, Rate: {instrument_data.get('avg_rate', 'N/A')}")
                row += 1
        
        if analytics.flow_metrics:
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="=== FLOW METRICS ===")
            row += 1
            for action, count in analytics.flow_metrics.action_breakdown.items():
                self.analytics_sheet.cell(row=row, column=1, value=f"Action {action}")
                self.analytics_sheet.cell(row=row, column=2, value=count)
                row += 1
        
        if analytics.risk_metrics:
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="=== RISK METRICS ===")
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Total DV01")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.risk_metrics.total_dv01)
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Concentration HHI")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.risk_metrics.concentration_hhi)
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Top 5 Concentration %")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.risk_metrics.top5_concentration)
            row += 1
        
        if analytics.realtime_metrics:
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="=== REAL-TIME METRICS ===")
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Liquidity Score")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.realtime_metrics.liquidity_score)
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Volume Last 5min")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.realtime_metrics.volume_last_5min)
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="Trades Last 5min")
            self.analytics_sheet.cell(row=row, column=2, value=analytics.realtime_metrics.trades_last_5min)
            row += 1
        
        if analytics.currency_metrics:
            row += 1
            self.analytics_sheet.cell(row=row, column=1, value="=== CURRENCY METRICS ===")
            row += 1
            for currency_data in analytics.currency_metrics.currency_breakdown:
                self.analytics_sheet.cell(row=row, column=1, value=f"Currency {currency_data['currency']}")
                self.analytics_sheet.cell(row=row, column=2, value=f"Notional: {currency_data['notional']}, Count: {currency_data['count']}")
                row += 1
    
    def load_trades_from_excel(self) -> List[Trade]:
        """Load all trades from today's Excel file."""
        trades = []
        try:
            self._ensure_file_exists()
            if not self.trades_sheet or self.trades_sheet.max_row < 2:
                logger.info("No trades found in Excel file")
                return trades
            
            # Read all rows (skip header row 1)
            for row in range(2, self.trades_sheet.max_row + 1):
                try:
                    trade_id = self.trades_sheet.cell(row=row, column=1).value
                    if not trade_id:
                        continue
                    
                    timestamp_str = self.trades_sheet.cell(row=row, column=2).value
                    if isinstance(timestamp_str, str):
                        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                    else:
                        timestamp = timestamp_str if isinstance(timestamp_str, datetime) else datetime.utcnow()
                    
                    # Reconstruct trade from Excel row
                    trade = Trade(
                        dissemination_identifier=str(trade_id),
                        action_type=self.trades_sheet.cell(row=row, column=3).value or "NEWT",
                        event_type="TRADE",
                        event_timestamp=timestamp,
                        execution_timestamp=timestamp,
                        effective_date=self.trades_sheet.cell(row=row, column=20).value if self.trades_sheet.max_column >= 20 else None,
                        expiration_date=self.trades_sheet.cell(row=row, column=13).value or None,
                        notional_amount_leg1=float(self.trades_sheet.cell(row=row, column=5).value or 0),
                        notional_amount_leg2=float(self.trades_sheet.cell(row=row, column=6).value or 0),
                        notional_currency_leg1=self.trades_sheet.cell(row=row, column=7).value or "",
                        notional_currency_leg2=self.trades_sheet.cell(row=row, column=8).value or "",
                        fixed_rate_leg1=self.trades_sheet.cell(row=row, column=9).value,
                        fixed_rate_leg2=self.trades_sheet.cell(row=row, column=10).value,
                        spread_leg1=None,
                        spread_leg2=self.trades_sheet.cell(row=row, column=11).value,
                        unique_product_identifier="",
                        unique_product_identifier_underlier_name=self.trades_sheet.cell(row=row, column=4).value or None,
                        platform_identifier=self.trades_sheet.cell(row=row, column=15).value or None,
                        package_indicator=self.trades_sheet.cell(row=row, column=17).value == "Yes" if self.trades_sheet.cell(row=row, column=17).value else False,
                        package_transaction_price=None,  # Not stored in Excel currently
                        strategy_id=self.trades_sheet.cell(row=row, column=16).value or None,
                        notional_eur=float(self.trades_sheet.cell(row=row, column=18).value or 0),
                        instrument=self.trades_sheet.cell(row=row, column=14).value or None,
                        is_forward=self.trades_sheet.cell(row=row, column=19).value == "Yes" if self.trades_sheet.max_column >= 19 and self.trades_sheet.cell(row=row, column=19).value else False,
                        effective_date_dt=None
                    )
                    trades.append(trade)
                except Exception as e:
                    logger.warning(f"Error loading trade from row {row}: {e}")
                    continue
            
            logger.info(f"Loaded {len(trades)} trades from Excel file")
        except Exception as e:
            logger.error(f"Error loading trades from Excel: {e}", exc_info=True)
        
        return trades

